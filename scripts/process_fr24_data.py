#!/usr/bin/env python3
"""
Process Flightradar24 APAC Busiest Routes Excel into JSON for Datavoy.

Reads: /Users/shiyuantian/Downloads/FR24_APAC_Busiest_Routes.xlsx
Writes:
    - data/fr24_apac_hubs.json
    - data/fr24_apac_routes.json

Outputs include a 0-100 Hub Connectivity Score based on:
    departures/week, destinations served, countries served, distinct airlines.
"""

import json
import math
from pathlib import Path
from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT_DIR / "data"
EXCEL_FILE = Path("/Users/shiyuantian/Downloads/FR24_APAC_Busiest_Routes.xlsx")

HUBS_OUT = DATA_DIR / "fr24_apac_hubs.json"
ROUTES_OUT = DATA_DIR / "fr24_apac_routes.json"


def normalize(values):
    """Min-max normalize a list of numbers to 0-1; return list of same length."""
    min_v = min(values)
    max_v = max(values)
    rng = max_v - min_v
    if rng == 0:
        return [0.0 for _ in values]
    return [(v - min_v) / rng for v in values]


def parse_hub_summary(ws):
    """Parse the 'Hub Summary' sheet."""
    rows = list(ws.iter_rows(values_only=True))
    # Header is row 5 (index 4)
    header = [c.strip() if isinstance(c, str) else c for c in rows[4]]
    records = []
    for row in rows[5:]:
        if not row or not row[1]:
            continue
        rec = dict(zip(header, row))
        iata = rec.get("IATA")
        # Skip total/summary rows and ensure valid 3-letter IATA code
        if not iata or not isinstance(iata, str) or len(iata) != 3 or iata.upper() == "TOTAL":
            continue
        records.append(rec)
    return records


def parse_all_routes(ws):
    """Parse the 'All Routes' sheet."""
    rows = list(ws.iter_rows(values_only=True))
    header = [c.strip() if isinstance(c, str) else c for c in rows[4]]
    records = []
    for row in rows[5:]:
        if not row or not row[1]:
            continue
        rec = dict(zip(header, row))
        if not rec.get("Origin"):
            continue
        records.append(rec)
    return records


def clean_hub(raw):
    """Convert raw hub summary row into clean dict."""
    return {
        "iata": raw.get("IATA"),
        "airport": raw.get("Airport"),
        "country": raw.get("Country"),
        "destinations_listed": int(raw.get("Destinations listed", 0) or 0),
        "destinations_with_service": int(raw.get("Destinations with service", 0) or 0),
        "countries_served": int(raw.get("Countries served", 0) or 0),
        "distinct_airlines": int(raw.get("Distinct airlines", 0) or 0),
        "departures_per_week": float(raw.get("Departures per week", 0) or 0),
        "busiest_1": raw.get("Busiest #1"),
        "busiest_2": raw.get("Busiest #2"),
        "busiest_3": raw.get("Busiest #3"),
    }


def clean_route(raw):
    """Convert raw route row into clean dict."""
    return {
        "origin": raw.get("Origin"),
        "origin_airport": raw.get("Origin airport"),
        "rank_in_hub": int(raw.get("Rank in hub", 0) or 0),
        "dest": raw.get("Dest"),
        "dest_icao": raw.get("Dest ICAO"),
        "destination_airport": raw.get("Destination airport"),
        "destination_city": raw.get("Destination city"),
        "country": raw.get("Country"),
        "departures_per_week": float(raw.get("Departures/wk", 0) or 0),
        "arrivals_per_week": float(raw.get("Arrivals/wk", 0) or 0),
        "total_per_week": float(raw.get("Total/wk", 0) or 0),
        "flight_numbers": int(raw.get("Flight numbers", 0) or 0),
        "airlines_count": int(raw.get("Airlines", 0) or 0),
        "lead_airline": raw.get("Lead airline"),
        "airlines_operating": raw.get("Airlines operating"),
        "common_aircraft": raw.get("Common aircraft"),
        "legs_observed": int(raw.get("Legs observed", 0) or 0),
        "days_observed": int(raw.get("Days observed", 0) or 0),
        "latitude": raw.get("Latitude"),
        "longitude": raw.get("Longitude"),
    }


def add_connectivity_scores(hubs):
    """Add a 0-100 connectivity score to each hub."""
    dep = normalize([h["departures_per_week"] for h in hubs])
    dst = normalize([h["destinations_with_service"] for h in hubs])
    ctr = normalize([h["countries_served"] for h in hubs])
    aln = normalize([h["distinct_airlines"] for h in hubs])

    for i, h in enumerate(hubs):
        score = (
            0.40 * dep[i]
            + 0.30 * dst[i]
            + 0.20 * ctr[i]
            + 0.10 * aln[i]
        ) * 100
        h["connectivity_score"] = round(score, 1)
    return hubs


def main():
    print(f"Reading {EXCEL_FILE}")
    wb = load_workbook(EXCEL_FILE, data_only=True)

    hub_ws = wb["Hub Summary"]
    route_ws = wb["All Routes"]

    raw_hubs = parse_hub_summary(hub_ws)
    hubs = [clean_hub(h) for h in raw_hubs]
    hubs = add_connectivity_scores(hubs)
    # Sort by connectivity score descending
    hubs.sort(key=lambda x: x["connectivity_score"], reverse=True)

    raw_routes = parse_all_routes(route_ws)
    routes = [clean_route(r) for r in raw_routes]

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    HUBS_OUT.write_text(json.dumps(hubs, ensure_ascii=False, indent=2))
    ROUTES_OUT.write_text(json.dumps(routes, ensure_ascii=False, indent=2))

    print(f"Wrote {HUBS_OUT} ({len(hubs)} hubs)")
    print(f"Wrote {ROUTES_OUT} ({len(routes)} routes)")

    # Print top 5
    print("\nTop 5 hubs by connectivity score:")
    for h in hubs[:5]:
        print(f"  {h['iata']:>3}  {h['airport']:<30}  score={h['connectivity_score']}")


if __name__ == "__main__":
    main()
